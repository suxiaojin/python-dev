'''
办公自动化--python操作excel
    openpyxl
    from openpyxl import Workbook
    #写入方法

    wb=Workbook()    --创建对象
    sheet=wb.active  --获取当前sheet
    wb.create_sheet(title=None,index=None) --创建新的sheet
    wb.remove(sheet) ---删除sheet
    sheet['A1]=42  ---单元格设置值
    sheet.append[1,2,3] ---插入一行数据
    wb.save(fpath) ---保存数据

    #读取方法
    wb=load_workbook('xxx.xlsx')  ---导入数据
    wb.sheetnames ---获取所有sheet名称
    sheet=wb[sheetname]  --获取sheet
    cell-sheet['A1'] ---获取一个单元格
    cell.value ---获取单元格数据

'''

'''
发送邮件
'''
import smtplib
from email.mime.text import MIMEText
from email.header import Header
sender='xxx@126.com'
passwd='xxx'
to_addrs=['xxx','abc']
#邮件内容：
message=MIMEText('python邮件测试','plain','utf-8')
#邮件主题
subject='python邮件发送测试'
message['Subject']=Header(subject,'utf-8')
#创建smtp对象
smtp=smtplib.SMTP()
#连接163邮箱
smtp.connect('smtp.163.com',25)
#登录邮箱
smtp.login(sender,passwd)
#发送邮件
smtp.sendmail(sender,to_addrs,message.as_string())
smtp.quit()

'''
系统资源获取--psutil

cpu:
    psutil.cpu_times  ---cpu使用情况
    psutil.cpu_count() ---cpu数量
    psutil.cpu_percent() --- cpu使用率
    psutil.stats() ---- cpu状态
    
内存：
    psutil.virtual_memory() --- 获取内存相关信息
    m=psutil.virtual_memory()
    m.total -- 总内存
    m.available --- 可用内存
    m.used ---使用的内存
    m.free ---剩余内存
    m.buffers
    
磁盘：
    psutil.disk_partitions  --分区信息
    d=psutil.disk_usage('/') --磁盘使用情况
    d.total --磁盘大小
    d.used --使用磁盘
    d.free --空余磁盘
    d.percent -- 使用率
    
进程信息：
    psutil.pids() ---获取进程ID列表
    for pid in psutil.pids():
        p=psutil.Process(pid) ---获取进程信息
        p.name --进程名
        p.status ---当前状态
        p.cpu_time() --进程cpu主要信息
        
网络信息：
    psutil.net_if_addrs ---- 获取网卡信息
    psutil.net_if_stats() ---获取网卡状态
    
'''

'''
时间:
    #当前时间转换为时分秒
    time.strftime("%Y-%m-%d %H-%M-%S")

datetime:
    date+time
    datetime.datetime.now() ---获取当前时间

日历操作：
    calendar
'''

'''
定时任务四种实现方式
time.sleep
    需求：检查网络使用情况，每5秒统计一次
    
threading.Timer   

sched模块

APScheduler---定时调度与间隔调度
from apscheduler.schedulers.blocking import BlockingScheduler
        scheduler=BlockingScheduler
       
        #每隔固定时间执行tick_task任务
        scheduler.add_job(函数(tick_task)，’interval',seconds=5)
        #定时任务：每天
        scheduler.add_job(time_task,'cron',hour=9,minute=19)
        scheduler.start()
触发器 interval ---每隔多长时间触发
        data ----特定时间触发--只运行一次
        cron-----固定时间触发 
        
'''
import time
import psutil

#获取网络发送与接收的字节数
def get_network_io_bytes():
    packages=psutil.net_io_counters()
    return packages.bytes_sent,packages.bytes_recv

def process_network_speed(n):
    send,recv=get_network_io_bytes()
    ms=1024*1024/n
    while True:
        time.sleep(n)
        new_send,new_recv=get_network_io_bytes()
        sendbytes=new_send-send
        recvbytes=new_recv-recv
        print('send speed=%.3f m/s,recv speed=%.sf M/S' %(sendbytes /ms,recvbytes /ms))
        send,recv=new_send,new_recv


#每隔5秒执行
process_network_speed(5)


'''
paramiko包含2个核心组件 
    SSHClient
        常用方法:
            connect():实现远程服务器的连接与认证，对于该方法只有hostname是必传参数。
                常用参数
                    hostname 连接的目标主机
                    port=SSH_PORT 指定端口
                    username=None 验证的用户名
                    password=None 验证的用户密码
                    pkey=None 私钥方式用于身份验证
                    key_filename=None 一个文件名或文件列表，指定私钥文件
                    timeout=None 可选的tcp连接超时时间
                    allow_agent=True, 是否允许连接到ssh代理，默认为True 允许
                    look_for_keys=True 是否在~/.ssh中搜索私钥文件，默认为True 允许
                    compress=False, 是否打开压缩
            set_missing_host_key_policy()：设置远程服务器没有在know_hosts文件中记录时的应对策略。目前支持三种策略：
                AutoAddPolicy 自动添加主机名及主机密钥到本地HostKeys对象，不依赖load_system_host_key的配置。即新建立ssh连接时不需要再输入yes或no进行确认
            exec_command()：在远程服务器执行Linux命令的方法。
 举例：  
 import paramiko
 
   # 实例化SSHClient
   client = paramiko.SSHClient()
 
   # 自动添加策略，保存服务器的主机名和密钥信息，如果不添加，那么不再本地know_hosts文件中记录的主机将无法连接
   client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
 
   # 连接SSH服务端，以用户名和密码进行认证
   client.connect(hostname='192.168.1.105', port=22, username='root', password='123456')
 
   # 打开一个Channel并执行命令
   stdin, stdout, stderr = client.exec_command('df -h ')  # stdout 为正确输出，stderr为错误输出，同时是有1个变量有值
 
   # 打印执行结果
   print(stdout.read().decode('utf-8'))
   
   # 关闭SSHClient
   client.close()


    SFTPClient

'''
